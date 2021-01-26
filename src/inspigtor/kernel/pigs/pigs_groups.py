import collections
import logging

import openpyxl

import numpy as np

import pandas as pd

import scipy.stats as stats

import scikit_posthocs as sk

from inspigtor.kernel.pigs.pigs_pool import PigsPoolError
from inspigtor.kernel.utils.stats import statistical_functions
from inspigtor.kernel.utils.progress_bar import progress_bar


class PigsGroupsError(Exception):
    """
    """


class PigsGroups:
    """This class implements the groups of pigs. Each member of the groups is a PigsPool object.
    """

    def __init__(self):
        """Constructor
        """

        self._groups = collections.OrderedDict()

    def __contains__(self, group):

        return group in self._groups

    def __len__(self):

        return len(self._groups)

    def add_group(self, group, pigs_pool):
        """Add a new group.

        Args:
            group (str): the name of the group
            pigs_pool (inspigtor.kernel.pigs.pigs_pool.PigsPool): the pigs pool
        """

        if group in self._groups:
            logging.warning('The group {} has already been registered'.format(group))
            return

        self._groups[group] = pigs_pool

    def evaluate_global_group_effect(self, selected_property='APs', selected_groups=None, interval_indexes=None):
        """Performs a statistical test to check whether the selected groups belongs to the same distribution.
        If there are only two groups, a Mann-Whitney test is performed otherwise a Kruskal-Wallis test
        is performed.

        Args:
            selected_property (str): the selected property
            selected_groups (list of str): the selected group. if None all the groups of the model will be used.
            interval_indexes (list of int): the indexes of the record intervals to select. If None, all the
            record intervals will be used.

        Returns:clear
        Raises:
            PigsGroupsError: if the number of groups is less than two.
        """

        # If selected groups is not provided by the user take all the groups
        if selected_groups is None:
            selected_groups = list(self._groups.keys())
        else:
            all_groups = set(self._groups.keys())
            selected_groups = list(all_groups.intersection(selected_groups))

        if len(selected_groups) < 2:
            raise PigsGroupsError('There is less than two groups. Can not perform any global statistical test.')

        longest_timeline = []

        progress_bar.reset(len(selected_groups))
        averages_per_group = collections.OrderedDict()
        for i, group in enumerate(selected_groups):
            try:
                timeline, averages_per_group[group] = self._groups[group].get_statistics(
                    selected_property, selected_statistics='mean', interval_indexes=interval_indexes)
            except PigsPoolError as error:
                raise PigsGroupsError from error
            else:
                if len(timeline) > len(longest_timeline):
                    longest_timeline = timeline

                progress_bar.update(i+1)

        if not averages_per_group:
            raise PigsGroupsError('There is less than two groups. Can not perform any global statistical test.')

        progress_bar.reset(len(longest_timeline))

        p_values_per_time = []
        # Loop over the intervals
        for i, time in enumerate(longest_timeline):
            groups = []
            n_values_per_group = []
            uncomplete_group = False
            for averages in averages_per_group.values():
                # This interval is not defined for this group, skip the group
                if i >= averages.shape[0]:
                    uncomplete_group = True
                    n_values_per_group.append(0)
                else:
                    values = [v for v in averages[i, :] if not np.isnan(v)]
                    if not values:
                        uncomplete_group = True
                        n_values_per_group.append(0)
                    else:
                        groups.append(values)
                        n_values_per_group.append(len(values))

            if uncomplete_group:
                p_value = np.nan
            else:
                n_groups = len(groups)
                if n_groups < 2:
                    p_value = np.nan
                else:
                    if n_groups == 2:
                        p_value = stats.mannwhitneyu(*groups, alternative='two-sided').pvalue
                    else:
                        p_value = stats.kruskal(*groups).pvalue

            p_values_per_time.append(n_values_per_group + [p_value])

            progress_bar.update(i+1)

        group_names = list(averages_per_group.keys())
        columns = group_names + ['p value']
        p_values = pd.DataFrame(p_values_per_time, index=longest_timeline, columns=columns)

        return p_values

    def evaluate_global_time_effect(self, selected_property='APs', selected_groups=None, interval_indexes=None):
        """Performs a Friedman statistical test to check whether the averages defined for each pig over record intervals
        belongs to the same distribution.

        Args:
            selected_property (str): the selected property
            selected_groups (list of str): the selected group. if None all the groups will be used.
            interval_indexes (list of int): the indexes of the record intervals to select. If None, all the record intervals will be used.

        Returns:
            list of str: the valid groups for which the Friedman test could be performed.
            list of float: the Friedman p values
        """

        # If selected groups is not provided by the user take all the groups
        if selected_groups is None:
            selected_groups = list(self._groups.keys())
        else:
            all_groups = set(self._groups.keys())
            selected_groups = list(all_groups.intersection(selected_groups))

        progress_bar.reset(len(selected_groups))
        valid_groups = []
        p_values = []
        for i, group in enumerate(selected_groups):
            try:
                _, p_value = self._groups[group].evaluate_global_time_effect(selected_property=selected_property, interval_indexes=interval_indexes)
            except PigsPoolError:
                logging.error('Can not evaluate global time effect for group {}'.format(group))
                continue
            else:
                valid_groups.append(group)
                p_values.append(p_value)
            finally:
                progress_bar.update(i+1)

        return valid_groups, p_values

    def evaluate_pairwise_group_effect(self, selected_property='APs', selected_groups=None, interval_indexes=None):
        """Performs a pairwise statistical test to check whether each pair of groups belongs to the same distribution.
        This should be evaluated only if the number of groups is >= 2.

        Args:
            selected_property (str): the selected property.
            selected_groups (list of str): the selected group. if None all the groups of the model will be used.
            interval_indexes (list of int): the indexes of the record intervals to select. If None, all the
            record intervals will be used.

        Returns:
            list of int: the list of valid intervals for which the Dunn test could be performed
            dict: the p values defined for each pair of groups resulting from the Dunn test.

        Raises:
            PigsGroupsError: if the number of groups is less than two.
        """

        # If selected groups is not provided by the user take all the groups
        if selected_groups is None:
            selected_groups = list(self._groups.keys())
        else:
            all_groups = set(self._groups.keys())
            selected_groups = list(all_groups.intersection(selected_groups))

        if len(selected_groups) < 2:
            raise PigsGroupsError('There is less than two groups. Can not perform any global statistical test.')

        longest_timeline = []
        averages_per_group = collections.OrderedDict()
        for i, group in enumerate(selected_groups):
            try:
                timeline, averages_per_group[group] = self._groups[group].get_statistics(
                    selected_property, selected_statistics='mean', interval_indexes=interval_indexes)
            except PigsPoolError as error:
                raise PigsGroupsError from error
            else:
                if len(timeline) > len(longest_timeline):
                    longest_timeline = timeline

        group_names = list(averages_per_group.keys())

        if not averages_per_group:
            raise PigsGroupsError('There is less than two groups. Can not perform any global statistical test.')

        progress_bar.reset(len(longest_timeline))
        p_values = collections.OrderedDict()
        # Loop over the intervals
        for i, time in enumerate(longest_timeline):
            uncomplete_group = False
            for averages in averages_per_group.values():
                # This interval is not defined for this group, skip the group
                if i >= averages.shape[0]:
                    uncomplete_group = True
                else:
                    values = [v for v in averages[i, :] if not np.isnan(v)]
                    if not values:
                        uncomplete_group = True
                    else:
                        groups.append(values)

            if uncomplete_group or len(groups) < 2:
                p_values[time] = pd.DataFrame(np.nan, index=group_names, columns=group_names)
            else:
                p_values[time] = pd.DataFrame(sk.posthoc_dunn(groups).to_numpy(), index=group_names, columns=group_names)

            progress_bar.update(i+1)

        return p_values

    def evaluate_pairwise_time_effect(self, selected_property='APs', selected_groups=None, interval_indexes=None):
        """Performs a Dunn statistical test to check whether the averages defined for each pig over record intervals belongs
        to the same distribution pairwisely.

        Args:
            selected_property (str): the selected property
            interval_indexes (list of int): the indexes of the record intervals to select. If None, all the record intervals will be used.

        Returns:
            list of str: the valid groups for which the Dunn test could be performed.
            list of pandas.DataFrame: the p values matrix resulting from Dunn test
        """

        # If selected groups is not provided by the user take all the groups
        if selected_groups is None:
            selected_groups = list(self._groups.keys())
        else:
            all_groups = set(self._groups.keys())
            selected_groups = list(all_groups.intersection(selected_groups))

        progress_bar.reset(len(selected_groups))
        valid_groups = collections.OrderedDict()
        for i, group in enumerate(selected_groups):
            try:
                p_values = self._groups[group].evaluate_pairwise_time_effect(
                    selected_property=selected_property, interval_indexes=interval_indexes)
            except PigsPoolError:
                logging.error('Can not evaluate pairwise time effect for group {}'.format(group))
            else:
                valid_groups[group] = p_values
            finally:
                progress_bar.update(i+1)

        return valid_groups

    def export_statistics(self, filename, selected_property='APs', selected_groups=None, interval_indexes=None):
        """Export basic statistics (average, median, std, quartile ...) for each group and interval to an excel file.

        Args:
            filename (str): the output excel filename
            selected_property (str): the selected property
            selected_groups (list of str): the selected group. if None all the groups of the model will be used.
            interval_indexes (list of int): the indexes of the record intervals to select. If None, all the
            record intervals will be used.
        """

        if selected_groups is None:
            selected_groups = list(self._groups.keys())
        else:
            all_groups = set(self._groups.keys())
            selected_groups = list(all_groups.intersection(selected_groups))

        workbook = openpyxl.Workbook()
        # Remove the first empty sheet created by default
        workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))

        for group in selected_groups:

            try:
                reduced_averages = self._groups[group].reduced_statistics(selected_property, selected_statistics='mean',
                                                                          interval_indexes=interval_indexes, output_statistics=statistical_functions.keys())
            except PigsPoolError as error:
                logging.error(str(error))
                return

            # Create the excel worksheet
            workbook.create_sheet(group)
            worksheet = workbook.get_sheet_by_name(group)

            worksheet.cell(row=1, column=1).value = 'time'
            worksheet.cell(row=1, column=12).value = 'selected property'
            worksheet.cell(row=2, column=12).value = selected_property
            worksheet.cell(row=1, column=14).value = 'pigs'

            longest_timeline = []
            for reader in self._groups[group].pigs.values():
                timeline = reader.timeline
                if len(timeline) > len(longest_timeline):
                    longest_timeline = timeline

            # Add titles
            for col, func in enumerate(statistical_functions.keys()):
                worksheet.cell(row=1, column=col+2).value = func

                for row, value in enumerate(reduced_averages[func]):
                    worksheet.cell(row=row+2, column=1).value = longest_timeline[row]
                    worksheet.cell(row=row+2, column=col+2).value = value

            pigs = self._groups[group].pigs.keys()
            for row, pig in enumerate(pigs):
                worksheet.cell(row=row+2, column=14).value = pig

        try:
            workbook.save(filename)
        except PermissionError as error:
            logging.error(str(error))
            return

        logging.info('Exported successfully groups statistics in {} file'.format(filename))

    def get_group(self, group):

        return self._groups.get(group, None)

    @ property
    def groups(self):
        """Getter for self._groups attribute. Returns the registered groups.

        Returns:
            dict: the registered groups
        """

        return self._groups

    def premortem_statistics(self, n_last_intervals, selected_property='APs', selected_groups=None):
        """
        """

        if selected_groups is None:
            selected_groups = list(self._groups.keys())
        else:
            all_groups = set(self._groups.keys())
            selected_groups = list(all_groups.intersection(selected_groups))

        progress_bar.reset(len(selected_groups))
        p_values = []
        for i, group in enumerate(selected_groups):
            p_value = self._groups[group].premortem_statistics(n_last_intervals, selected_property=selected_property)
            p_values.append(p_value)
            progress_bar.update(i+1)

        return dict(zip(selected_groups, p_values))

    def remove_reader(self, filename):
        """
        """

        for pool in self._groups.values():

            if not pool.has_reader(filename):
                continue

            pool.remove_reader(filename)

    def set_record_interval(self, interval):
        """Set the record interval.

        Args:
            interval (4-tuple): the record interval. 4-tuple of the form (start,end,record,offset).
        """

        for pool in self._groups.values():

            pool.set_record_interval(interval)


if __name__ == '__main__':

    import sys
    from inspigtor.kernel.pigs.pigs_pool import PigsPool

    pool1 = PigsPool()
    pool1.add_reader(sys.argv[1])
    pool1.add_reader(sys.argv[2])

    pool2 = PigsPool()
    pool2.add_reader(sys.argv[1])
    pool2.add_reader(sys.argv[2])

    groups = PigsGroups()
    groups.add_group('group1', pool1)
    groups.add_group('group2', pool2)
    groups.set_record_interval(('00:00:00', '01:00:00', 300, 30))
    print(groups.evaluate_global_group_effect('APs'))
    print(groups.evaluate_pairwise_group_effect('APs'))
    groups.export_statistics('test.xlsx', selected_property='APs')
